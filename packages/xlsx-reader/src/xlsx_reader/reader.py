"""Core XLSX reading functionality."""

from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


def read_xlsx(file_path: str | Path) -> Dict[str, List[List[Any]]]:
    """
    Extract all data from an Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary mapping sheet names to their data (list of rows)

    Raises:
        FileNotFoundError: If the Excel file doesn't exist
        ValueError: If the file is not a valid Excel file
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        workbook = load_workbook(path, data_only=True)
        result = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = []

            for row in sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if any(cell is not None for cell in row):
                    data.append(list(row))

            result[sheet_name] = data

        return result
    except Exception as e:
        raise ValueError(f"Failed to read Excel file: {e}") from e


def extract_sheets(file_path: str | Path) -> List[Dict[str, Any]]:
    """
    Extract data from an Excel file, organized by sheet with metadata.

    Args:
        file_path: Path to the Excel file

    Returns:
        List of dictionaries with sheet information and data

    Raises:
        FileNotFoundError: If the Excel file doesn't exist
        ValueError: If the file is not a valid Excel file
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        workbook = load_workbook(path, data_only=True)
        sheets = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = []

            for row in sheet.iter_rows(values_only=True):
                # Skip completely empty rows
                if any(cell is not None for cell in row):
                    data.append(list(row))

            sheets.append(
                {
                    "name": sheet_name,
                    "rows": len(data),
                    "columns": len(data[0]) if data else 0,
                    "data": data,
                }
            )

        return sheets
    except Exception as e:
        raise ValueError(f"Failed to read Excel file: {e}") from e
